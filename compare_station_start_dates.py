import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
import tkinter as tk
from tkinter import filedialog, messagebox

# ---------------- CONFIG ----------------

MASTER_XLSX = "station_start_date_summary.xlsx"
HEADER_LINES = 4

TS_FORMATS = [
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
]

TABLE_TYPES = [
    "TableDay",
    "TableETHour",
    "TableHour",
    "SYNOP",
    "Table10m",
    "TableSolarCharger10m",
]

# ---------------- HELPERS ----------------

def parse_ts(text):
    text = text.strip().strip('"')
    for fmt in TS_FORMATS:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def get_start_end(path):
    with open(path, "r", errors="ignore") as f:
        lines = f.readlines()

    start = None
    end = None

    # Start timestamp (after header)
    for ln in lines[HEADER_LINES:]:
        if not ln.strip():
            continue
        ts = parse_ts(ln.split(",")[0])
        if ts:
            start = ts
            break

    # End timestamp (scan backwards)
    for ln in reversed(lines):
        if not ln.strip():
            continue
        ts = parse_ts(ln.split(",")[0])
        if ts:
            end = ts
            break

    return start, end


def get_earliest_station_start(folder):
    earliest = None

    for fname in os.listdir(folder):
        if not fname.endswith(".dat"):
            continue

        path = os.path.join(folder, fname)
        start, _ = get_start_end(path)

        if start:
            if earliest is None or start < earliest:
                earliest = start

    return earliest

# ---------------- MAIN ----------------

def main():
    root = tk.Tk()
    root.withdraw()

    folder_a = filedialog.askdirectory(title="Select FIRST station folder")
    if not folder_a:
        return

    folder_b = filedialog.askdirectory(title="Select SECOND station folder")
    if not folder_b:
        return

    station_a = os.path.basename(os.path.normpath(folder_a))
    station_b = os.path.basename(os.path.normpath(folder_b))

    start_a = get_earliest_station_start(folder_a)
    start_b = get_earliest_station_start(folder_b)

    if not start_a or not start_b:
        messagebox.showerror(
            "Error",
            "Could not determine start dates for one or both stations."
        )
        return

    if start_a < start_b:
        winner = station_a
    elif start_b < start_a:
        winner = station_b
    else:
        winner = "Same start date"

    # Load or create workbook
    if os.path.exists(MASTER_XLSX):
        wb = load_workbook(MASTER_XLSX)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    sheet_name = "Station_Comparisons"

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        ws.append([
            "Station A",
            "Station B",
            "Earliest Start A",
            "Earliest Start B",
            "Started Earlier",
        ])

    ws.append([
        station_a,
        station_b,
        start_a,
        start_b,
        winner,
    ])

    wb.save(MASTER_XLSX)

    messagebox.showinfo(
        "Comparison Complete",
        f"Comparison saved to:\n{MASTER_XLSX}\n\n"
        f"{station_a}: {start_a}\n"
        f"{station_b}: {start_b}\n\n"
        f"Started earlier: {winner}"
    )

    print("✅ Comparison saved")
    print(f"{station_a} → {start_a}")
    print(f"{station_b} → {start_b}")
    print(f"Started earlier → {winner}")


if __name__ == "__main__":
    main()
