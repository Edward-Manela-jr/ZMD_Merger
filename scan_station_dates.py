# import os
# import argparse
# from datetime import datetime
# from openpyxl import Workbook

# TS_FORMATS = [
#     "%Y-%m-%d %H:%M:%S",
#     "%Y-%m-%d %H:%M",
# ]

# HEADER_LINES = 4

# TABLE_TYPES = [
#     "TableDay",
#     "TableETHour",
#     "TableHour",
#     "SYNOP",
#     "Table10m",
#     "TableSolarCharger10m",
# ]


# def parse_ts(text):
#     text = text.strip().strip('"')
#     for fmt in TS_FORMATS:
#         try:
#             return datetime.strptime(text, fmt)
#         except ValueError:
#             pass
#     return None


# def get_start_end(path):
#     with open(path, "r", errors="ignore") as f:
#         lines = f.readlines()

#     start = None
#     end = None

#     # start timestamp (skip header)
#     for ln in lines[HEADER_LINES:]:
#         if not ln.strip():
#             continue
#         ts = parse_ts(ln.split(",")[0])
#         if ts:
#             start = ts
#             break

#     # end timestamp (scan backwards)
#     for ln in reversed(lines):
#         if not ln.strip():
#             continue
#         ts = parse_ts(ln.split(",")[0])
#         if ts:
#             end = ts
#             break

#     return start, end


# def detect_table_type(name):
#     for t in TABLE_TYPES:
#         if t in name:
#             return t
#     return "Unknown"


# def main():
#     p = argparse.ArgumentParser()
#     p.add_argument("--src", required=True, help="Station folder")
#     p.add_argument(
#         "--out",
#         default="station_date_summary.xlsx",
#         help="Excel output file",
#     )
#     args = p.parse_args()

#     files = [f for f in os.listdir(args.src) if f.endswith(".dat")]

#     if not files:
#         print("❌ No .dat files found")
#         return

#     station_name = os.path.basename(os.path.normpath(args.src))

#     wb = Workbook()
#     ws = wb.active
#     ws.title = station_name

#     # Header row
#     ws.append([
#         "Station",
#         "File Name",
#         "Table Type",
#         "Start Date",
#         "End Date",
#     ])

#     print(f"\n📂 Scanning station: {station_name}\n")

#     for fname in sorted(files):
#         path = os.path.join(args.src, fname)
#         table_type = detect_table_type(fname)

#         start, end = get_start_end(path)

#         print(f"{table_type}")
#         print(f"  File : {fname}")

#         if start and end:
#             print(f"  Start: {start}")
#             print(f"  End  : {end}")
#         else:
#             print("  ❌ Could not detect timestamps")

#         print()

#         ws.append([
#             station_name,
#             fname,
#             table_type,
#             start,
#             end,
#         ])

#     wb.save(args.out)
#     print(f"✅ Excel file written → {args.out}")


# if __name__ == "__main__":
#     main()













import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
import tkinter as tk
from tkinter import filedialog, messagebox

# ---------------- CONFIG ----------------

MASTER_XLSX = "station_date_summary.xlsx"
HEADER_LINES = 4

TS_FORMATS = [
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
]

TABLE_TYPES = [
    "TableDay",
    "TableETHour",
    "TableHour",
    "SYNOP",
    "Table10m",
    "TableSolarCharger10m",
]

# ---------------- HELPERS ----------------

def parse_ts(text):
    text = text.strip().strip('"')
    for fmt in TS_FORMATS:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def get_start_end(path):
    with open(path, "r", errors="ignore") as f:
        lines = f.readlines()

    start = None
    end = None

    # Start timestamp (after header)
    for ln in lines[HEADER_LINES:]:
        if not ln.strip():
            continue
        ts = parse_ts(ln.split(",")[0])
        if ts:
            start = ts
            break

    # End timestamp (scan backwards)
    for ln in reversed(lines):
        if not ln.strip():
            continue
        ts = parse_ts(ln.split(",")[0])
        if ts:
            end = ts
            break

    return start, end


def detect_table_type(name):
    for t in TABLE_TYPES:
        if t in name:
            return t
    return "Unknown"

# ---------------- MAIN ----------------

def main():
    root = tk.Tk()
    root.withdraw()

    folder = filedialog.askdirectory(
        title="Select Station Folder (contains .dat files)"
    )

    if not folder:
        return

    files = [f for f in os.listdir(folder) if f.endswith(".dat")]

    if not files:
        messagebox.showerror("Error", "No .dat files found in selected folder.")
        return

    station = os.path.basename(os.path.normpath(folder))

    # Load or create master workbook
    if os.path.exists(MASTER_XLSX):
        wb = load_workbook(MASTER_XLSX)
    else:
        wb = Workbook()
        wb.remove(wb.active)  # remove default sheet

    # If station sheet exists → confirm overwrite
    if station in wb.sheetnames:
        overwrite = messagebox.askyesno(
            "Sheet Exists",
            f"A sheet for '{station}' already exists.\n\nOverwrite it?"
        )
        if not overwrite:
            messagebox.showinfo("Cancelled", "Operation cancelled.")
            return
        del wb[station]

    ws = wb.create_sheet(title=station)

    # Header row
    ws.append([
        "Station",
        "File Name",
        "Table Type",
        "Start Date",
        "End Date",
    ])

    print(f"\n📂 Scanning station: {station}\n")

    for fname in sorted(files):
        path = os.path.join(folder, fname)
        table_type = detect_table_type(fname)
        start, end = get_start_end(path)

        print(f"{table_type}")
        print(f"  File : {fname}")

        if start and end:
            print(f"  Start: {start}")
            print(f"  End  : {end}")
        else:
            print("  ❌ Could not detect timestamps")

        print()

        ws.append([
            station,
            fname,
            table_type,
            start,
            end,
        ])

    wb.save(MASTER_XLSX)

    messagebox.showinfo(
        "Done",
        f"Station '{station}' added to:\n{MASTER_XLSX}"
    )

    print(f"✅ Updated Excel → {MASTER_XLSX}")


if __name__ == "__main__":
    main()
